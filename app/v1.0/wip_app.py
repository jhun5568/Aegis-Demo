# WIP(Work-In-Process) 현황 관리 앱
# Project Aegis - WIP Management System v1.0
# 작성일: 2025.09.29

import streamlit as st
import pandas as pd
import numpy as np
from datetime import date, datetime, timedelta
import os
import io
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 페이지 설정
st.set_page_config(
    page_title="WIP 현황관리 시스템",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="expanded"
)

class WIPManager:
    """WIP(Work-In-Process) 현황 관리 시스템"""
    
    def __init__(self, database_path=None):
            if database_path:
                self.database_file = database_path
            else:
                # 가장 간단한 방법: 같은 폴더에서 찾기
                current_dir = os.path.dirname(os.path.abspath(__file__))
                self.database_file = os.path.join(current_dir, "material_database.xlsx")
                
                print(f"🔍 WIP 앱 초기화")
                print(f"현재 디렉토리: {current_dir}")
                print(f"찾는 파일: {self.database_file}")
                print(f"파일 존재 여부: {os.path.exists(self.database_file)}")
                
                if not os.path.exists(self.database_file):
                    print(f"⚠️ 파일이 없으므로 새로 생성할 예정: {self.database_file}")
            
            self.stages = ["Cut", "Bend", "Laser", "Paint", "QA", "Receive"]
            self.stage_colors = {
                "Cut": "#FF6B6B",
                "Bend": "#4ECDC4", 
                "Laser": "#45B7D1",
                "Paint": "#96CEB4",
                "QA": "#FECA57",
                "Receive": "#6C5CE7"
            }        
    def _ukey(self, scope, *parts):
        """고유 키 생성기"""
        import re
        norm = [re.sub(r'[^0-9A-Za-z]+', '_', str(p)) for p in parts if p is not None]
        return f"wip_{scope}_" + "_".join(norm)
    
    def _safe_float(self, x, default=0.0):
        """안전한 float 변환"""
        try:
            v = float(x)
            if pd.isna(v):
                return default
            return v
        except Exception:
            return default
            
    def _safe_int(self, x, default=0):
        """안전한 int 변환"""
        try:
            return int(float(x))
        except Exception:
            return default

    def _ensure_wip_sheets(self):
            """WIP 관련 시트가 없으면 생성"""
            try:
                # 파일 경로가 None인지 먼저 확인
                if not self.database_file:
                    st.error("❌ 데이터베이스 파일 경로가 설정되지 않았습니다.")
                    return False
                
                print(f"📂 데이터베이스 파일 경로: {self.database_file}")
                
                if not os.path.exists(self.database_file):
                    print(f"📝 데이터베이스 파일이 없어서 새로 생성합니다: {self.database_file}")
                    # 데이터베이스 파일이 없으면 새로 생성
                    wb = Workbook()
                    wb.remove(wb.active)
                else:
                    print(f"📖 기존 데이터베이스 파일을 로드합니다: {self.database_file}")
                    wb = load_workbook(self.database_file)

                # 필요한 시트들과 헤더 정의
                required_sheets = {
                    "Orders": [
                        "order_id", "project", "vendor", "order_date", 
                        "due_date", "status", "memo", "created_at"
                    ],
                    "OrderItems": [
                        "order_id", "item", "spec", "unit", 
                        "qty_ordered", "qty_received", "notes"
                    ],
                    "ProcessEvents": [
                        "event_id", "order_id", "item", "stage", 
                        "planned_at", "done_at", "progress", "note", 
                        "created_at", "created_by"
                    ]
                }
                
                sheets_created = []
                for sheet_name, headers in required_sheets.items():
                    if sheet_name not in wb.sheetnames:
                        ws = wb.create_sheet(title=sheet_name)
                        ws.append(headers)
                        sheets_created.append(sheet_name)
                        print(f"✅ {sheet_name} 시트 생성됨")
                
                # 파일 저장 시 디렉토리가 없으면 생성
                os.makedirs(os.path.dirname(self.database_file), exist_ok=True)
                wb.save(self.database_file)
                
                if sheets_created:
                    st.info(f"WIP 시트 생성 완료: {', '.join(sheets_created)}")
                else:
                    print("✅ 모든 필요한 시트가 이미 존재합니다.")
                    
                return True
                
            except Exception as e:
                st.error(f"WIP 시트 생성 중 오류: {e}")
                print(f"❌ 상세 오류: {e}")
                return False
            
    def _load_wip_data(self):
        """WIP 관련 데이터 로딩"""
        try:
            # 시트 보장
            if not self._ensure_wip_sheets():
                return None, None, None
            
            # 데이터 로딩
            with pd.ExcelFile(self.database_file) as xls:
                # Orders 로딩
                try:
                    orders = pd.read_excel(xls, "Orders")
                except:
                    orders = pd.DataFrame(columns=[
                        "order_id", "project", "vendor", "order_date", 
                        "due_date", "status", "memo", "created_at"
                    ])
                
                # OrderItems 로딩  
                try:
                    items = pd.read_excel(xls, "OrderItems")
                except:
                    items = pd.DataFrame(columns=[
                        "order_id", "item", "spec", "unit", 
                        "qty_ordered", "qty_received", "notes"
                    ])
                
                # ProcessEvents 로딩
                try:
                    events = pd.read_excel(xls, "ProcessEvents")
                except:
                    events = pd.DataFrame(columns=[
                        "event_id", "order_id", "item", "stage", 
                        "planned_at", "done_at", "progress", "note", 
                        "created_at", "created_by"
                    ])
            
            # 데이터 타입 정리
            orders = self._clean_orders_data(orders)
            items = self._clean_items_data(items)  
            events = self._clean_events_data(events)
            
            return orders, items, events
            
        except Exception as e:
            st.error(f"데이터 로딩 중 오류: {e}")
            return None, None, None

    def _clean_orders_data(self, df):
        """Orders 데이터 정리"""
        if df.empty:
            return df
            
        # 날짜 컬럼 처리
        for col in ["order_date", "due_date"]:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors="coerce").dt.date
        
        if "created_at" in df.columns:
            df["created_at"] = pd.to_datetime(df["created_at"], errors="coerce")
            
        # 문자열 컬럼 정리
        for col in ["order_id", "project", "vendor", "status", "memo"]:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip()
                df[col] = df[col].replace('nan', '')
                
        return df

    def _clean_items_data(self, df):
        """OrderItems 데이터 정리"""
        if df.empty:
            return df
            
        # 문자열 컬럼 정리
        for col in ["order_id", "item", "spec", "unit", "notes"]:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip()
                df[col] = df[col].replace('nan', '')
        
        # 숫자 컬럼 정리
        for col in ["qty_ordered", "qty_received"]:
            if col in df.columns:
                df[col] = df[col].apply(lambda x: self._safe_float(x, 0.0))
                
        return df

    def _clean_events_data(self, df):
        """ProcessEvents 데이터 정리"""
        if df.empty:
            return df
            
        # 날짜/시간 컬럼 처리
        for col in ["planned_at", "done_at", "created_at"]:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors="coerce")
        
        # 문자열 컬럼 정리
        for col in ["event_id", "order_id", "item", "stage", "note", "created_by"]:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip()
                df[col] = df[col].replace('nan', '')
        
        # 진행률 컬럼 정리
        if "progress" in df.columns:
            df["progress"] = df["progress"].apply(lambda x: self._safe_int(x, 0))
            df["progress"] = df["progress"].clip(0, 100)
            
        return df
    
    def _calculate_progress(self, orders, events):
        """발주별 진행률 계산"""
        if orders.empty or events.empty:
            orders["progress_pct"] = 0
            orders["current_stage"] = "미시작"
            orders["stage_status"] = {}
            return orders
        
        # 최신 이벤트만 집계 (order_id + stage별 가장 최근)
        events_latest = (events.sort_values("created_at")
                        .drop_duplicates(subset=["order_id", "stage"], keep="last"))
        
        # 단계별 완료 판단 (done_at이 있거나 progress >= 100)
        events_latest["is_completed"] = (
            events_latest["done_at"].notna() | 
            (events_latest["progress"] >= 100)
        )
        
        def calculate_order_progress(order_id):
            """개별 발주의 진행률 계산"""
            order_events = events_latest[events_latest["order_id"] == str(order_id)]
            
            if order_events.empty:
                return 0, "미시작", {}
            
            # 각 단계별 상태 확인
            stage_status = {}
            completed_stages = 0
            current_stage = "미시작"
            
            for stage in self.stages:
                stage_events = order_events[order_events["stage"] == stage]
                
                if stage_events.empty:
                    stage_status[stage] = "대기"
                else:
                    latest_event = stage_events.iloc[-1]
                    if latest_event["is_completed"]:
                        stage_status[stage] = "완료"
                        completed_stages += 1
                    else:
                        stage_status[stage] = "진행중"
                        if current_stage == "미시작":
                            current_stage = stage
            
            # 모든 단계가 완료되지 않았다면 다음 대기 단계 찾기
            if current_stage == "미시작" and completed_stages > 0:
                for stage in self.stages:
                    if stage_status.get(stage) == "대기":
                        current_stage = stage
                        break
                if current_stage == "미시작":
                    current_stage = "완료"
            
            # 진행률 계산
            progress_pct = int((completed_stages / len(self.stages)) * 100)
            
            return progress_pct, current_stage, stage_status
        
        # 각 발주에 대해 진행률 계산
        progress_data = []
        for _, order in orders.iterrows():
            pct, stage, status = calculate_order_progress(order["order_id"])
            progress_data.append({
                "order_id": order["order_id"],
                "progress_pct": pct,
                "current_stage": stage,
                "stage_status": status
            })
        
        progress_df = pd.DataFrame(progress_data)
        
        # 원본 데이터와 병합
        orders = orders.merge(progress_df, on="order_id", how="left")
        orders["progress_pct"] = orders["progress_pct"].fillna(0).astype(int)
        orders["current_stage"] = orders["current_stage"].fillna("미시작")
        
        return orders

    def _render_dashboard_cards(self, orders):
        """대시보드 KPI 카드 렌더링"""
        if orders.empty:
            st.info("발주 데이터가 없습니다.")
            return
        
        # KPI 계산
        total_orders = len(orders)
        wip_orders = len(orders[orders["progress_pct"] < 100])
        completed_orders = len(orders[orders["progress_pct"] >= 100])
        
        # 지연 계산 (납기일이 지났는데 완료되지 않은 것)
        today = date.today()
        overdue_orders = len(orders[
            (orders["due_date"].notna()) & 
            (orders["due_date"] < today) & 
            (orders["progress_pct"] < 100)
        ])
        
        # 이번 주 완료 예정
        week_end = today + timedelta(days=7)
        thisweek_due = len(orders[
            (orders["due_date"].notna()) & 
            (orders["due_date"] <= week_end) & 
            (orders["due_date"] >= today) &
            (orders["progress_pct"] < 100)
        ])
        
        # 카드 렌더링
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric(
                "총 발주 건수", 
                f"{total_orders}",
                help="전체 발주 건수"
            )
        
        with col2:
            delta_color = "normal" if wip_orders == 0 else "off"
            st.metric(
                "진행 중", 
                f"{wip_orders}",
                delta=f"{wip_orders}/{total_orders}",
                delta_color=delta_color,
                help="진행 중인 발주 건수"
            )
            
        with col3:
            delta_color = "inverse" if overdue_orders > 0 else "normal"
            st.metric(
                "지연 건수", 
                f"{overdue_orders}",
                delta=f"긴급 처리 필요" if overdue_orders > 0 else "정상",
                delta_color=delta_color,
                help="납기일이 지난 미완료 발주"
            )
            
        with col4:
            st.metric(
                "이번주 완료 예정", 
                f"{thisweek_due}",
                help="이번 주 내 완료 예정 발주"
            )

    def _render_filters(self, orders):
        """필터 UI 렌더링"""
        st.subheader("🔍 필터")
        
        col1, col2, col3, col4 = st.columns([2, 2, 2, 2])
        
        with col1:
            projects = ["(전체)"] + sorted([
                str(x) for x in orders["project"].dropna().unique().tolist()
            ])
            f_project = st.selectbox(
                "프로젝트", 
                projects, 
                key=self._ukey("filter_project")
            )
        
        with col2:
            vendors = ["(전체)"] + sorted([
                str(x) for x in orders["vendor"].dropna().unique().tolist()
            ])
            f_vendor = st.selectbox(
                "업체", 
                vendors, 
                key=self._ukey("filter_vendor")
            )
            
        with col3:
            status_options = ["(전체)", "진행중", "완료", "지연"]
            f_status = st.selectbox(
                "상태", 
                status_options,
                key=self._ukey("filter_status")
            )
            
        with col4:
            f_stages = st.multiselect(
                "단계", 
                self.stages,
                key=self._ukey("filter_stages")
            )
        
        # 날짜 필터
        col5, col6 = st.columns(2)
        with col5:
            f_date_from = st.date_input(
                "납기 시작일", 
                value=date.today() - timedelta(days=30),
                key=self._ukey("filter_date_from")
            )
        with col6:
            f_date_to = st.date_input(
                "납기 종료일", 
                value=date.today() + timedelta(days=30),
                key=self._ukey("filter_date_to")
            )
        
        return {
            "project": f_project,
            "vendor": f_vendor, 
            "status": f_status,
            "stages": f_stages,
            "date_from": f_date_from,
            "date_to": f_date_to
        }

    def _apply_filters(self, orders, filters):
        """필터 적용"""
        filtered = orders.copy()
        
        # 프로젝트 필터
        if filters["project"] != "(전체)":
            filtered = filtered[filtered["project"] == filters["project"]]
        
        # 업체 필터
        if filters["vendor"] != "(전체)":
            filtered = filtered[filtered["vendor"] == filters["vendor"]]
            
        # 상태 필터
        today = date.today()
        if filters["status"] == "진행중":
            filtered = filtered[filtered["progress_pct"] < 100]
        elif filters["status"] == "완료":
            filtered = filtered[filtered["progress_pct"] >= 100]
        elif filters["status"] == "지연":
            filtered = filtered[
                (filtered["due_date"].notna()) & 
                (filtered["due_date"] < today) & 
                (filtered["progress_pct"] < 100)
            ]
        
        # 날짜 필터
        if filters["date_from"]:
            filtered = filtered[
                filtered["due_date"].fillna(date(2100, 1, 1)) >= filters["date_from"]
            ]
        if filters["date_to"]:
            filtered = filtered[
                filtered["due_date"].fillna(date(1900, 1, 1)) <= filters["date_to"]
            ]
            
        return filtered

    def _render_stage_chips(self, stage_status):
        """단계별 상태 칩 렌더링"""
        if not stage_status:
            return "미시작"
        
        chips = []
        for stage in self.stages:
            status = stage_status.get(stage, "대기")
            if status == "완료":
                chips.append(f"✅ {stage}")
            elif status == "진행중":
                chips.append(f"🟡 {stage}")
            else:
                chips.append(f"⚪ {stage}")
        
        return " | ".join(chips)

    def _render_progress_bar(self, progress_pct):
        """진행률 바 렌더링"""
        if progress_pct >= 100:
            color = "green"
        elif progress_pct >= 50:
            color = "orange"
        else:
            color = "red"
            
        return f"""
        <div style="background-color: #f0f0f0; border-radius: 10px; padding: 2px;">
            <div style="background-color: {color}; width: {progress_pct}%; 
                        height: 20px; border-radius: 8px; text-align: center; 
                        color: white; font-weight: bold; line-height: 20px;">
                {progress_pct}%
            </div>
        </div>
        """    
    def _render_main_table(self, orders, items):
        """메인 발주 테이블 렌더링"""
        if orders.empty:
            st.info("조건에 맞는 발주가 없습니다.")
            return None
            
        st.subheader("📋 발주 현황")
        
        # 표시용 데이터 준비
        display_data = []
        for _, order in orders.iterrows():
            # 납기일 표시 형식
            due_str = order["due_date"].strftime("%Y-%m-%d") if pd.notna(order["due_date"]) else "미정"
            
            # 지연 여부 체크
            today = date.today()
            is_overdue = (pd.notna(order["due_date"]) and 
                         order["due_date"] < today and 
                         order["progress_pct"] < 100)
            
            if is_overdue:
                due_str += " ⚠️"
            
            # 품목 수 계산
            order_items = items[items["order_id"] == order["order_id"]]
            item_count = len(order_items)
            
            display_data.append({
                "발주번호": order["order_id"],
                "프로젝트": order["project"],
                "업체": order["vendor"],
                "품목수": f"{item_count}개",
                "납기일": due_str,
                "진행률": f"{order['progress_pct']}%",
                "현재단계": order["current_stage"],
                "상태": "지연" if is_overdue else ("완료" if order["progress_pct"] >= 100 else "진행중")
            })
        
        if not display_data:
            st.info("표시할 데이터가 없습니다.")
            return None
        
        display_df = pd.DataFrame(display_data)
        
        # 테이블 스타일링을 위한 함수
        def style_row(row):
            if "⚠️" in str(row["납기일"]):
                return ["background-color: #ffebee"] * len(row)
            elif row["상태"] == "완료":
                return ["background-color: #e8f5e8"] * len(row)
            else:
                return [""] * len(row)
        
        styled_df = display_df.style.apply(style_row, axis=1)
        st.dataframe(styled_df, use_container_width=True)
        
        # 발주 선택
        st.markdown("---")
        order_options = [f"{row['발주번호']} - {row['프로젝트']}" for row in display_data]
        
        if order_options:
            selected_option = st.selectbox(
                "상세 보기할 발주 선택",
                ["선택하세요..."] + order_options,
                key=self._ukey("select_order")
            )
            
            if selected_option != "선택하세요...":
                selected_order_id = selected_option.split(" - ")[0]
                return selected_order_id
        
        return None

    def _render_detail_panel(self, order_id, orders, items, events):
        """선택된 발주의 상세 패널 렌더링"""
        # 발주 정보 가져오기
        order_info = orders[orders["order_id"] == order_id]
        if order_info.empty:
            st.error("발주 정보를 찾을 수 없습니다.")
            return
        
        order = order_info.iloc[0]
        
        # 헤더
        st.subheader(f"🗂️ {order['project']} - {order['vendor']}")
        st.caption(f"발주번호: {order['order_id']}")
        
        # 기본 정보
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric("진행률", f"{order['progress_pct']}%")
            
        with col2:
            due_date = order["due_date"].strftime("%Y-%m-%d") if pd.notna(order["due_date"]) else "미정"
            st.metric("납기일", due_date)
            
        with col3:
            st.metric("현재 단계", order["current_stage"])
        
        # 진행률 바 표시
        st.markdown("**진행 상황**")
        progress_html = self._render_progress_bar(order["progress_pct"])
        st.markdown(progress_html, unsafe_allow_html=True)
        
        # 단계별 상태
        if hasattr(order, 'stage_status') and order['stage_status']:
            st.markdown("**단계별 현황**")
            stage_chips = self._render_stage_chips(order['stage_status'])
            st.markdown(stage_chips)
        
        # 발주 품목 리스트
        order_items = items[items["order_id"] == order_id]
        if not order_items.empty:
            st.markdown("**발주 품목**")
            display_items = order_items[["item", "spec", "unit", "qty_ordered", "qty_received"]].copy()
            display_items.columns = ["품목", "규격", "단위", "발주수량", "입고수량"]
            st.dataframe(display_items, use_container_width=True)
        
        # 최근 이벤트 이력
        order_events = events[events["order_id"] == order_id].sort_values("created_at", ascending=False)
        if not order_events.empty:
            st.markdown("**최근 진행 이력**")
            with st.expander("이력 보기", expanded=False):
                for _, event in order_events.head(10).iterrows():
                    created_time = event["created_at"].strftime("%Y-%m-%d %H:%M") if pd.notna(event["created_at"]) else "미상"
                    done_time = event["done_at"].strftime("%Y-%m-%d") if pd.notna(event["done_at"]) else "진행중"
                    
                    st.write(f"**{event['stage']}** - {event['progress']}% | 완료: {done_time} | 등록: {created_time}")
                    if event["note"]:
                        st.caption(f"메모: {event['note']}")
                    st.divider()

    def _render_update_form(self, order_id, items):
        """이벤트 업데이트 폼 렌더링"""
        st.markdown("### 🔧 진행 상황 업데이트")
        
        with st.form(key=self._ukey("update_form", order_id)):
            col1, col2 = st.columns(2)
            
            with col1:
                stage = st.selectbox("단계", self.stages, key=self._ukey("stage"))
                progress = st.slider("진행률 (%)", 0, 100, 100, 5, key=self._ukey("progress"))
                
            with col2:
                done_date = st.date_input("완료일", value=date.today(), key=self._ukey("done_date"))
                
                # 품목 선택 (선택사항)
                order_items = items[items["order_id"] == order_id]
                item_options = ["전체 품목"] + order_items["item"].tolist() if not order_items.empty else ["전체 품목"]
                selected_item = st.selectbox("대상 품목", item_options, key=self._ukey("item"))
            
            note = st.text_area("메모", placeholder="진행 상황에 대한 메모를 입력하세요...", key=self._ukey("note"))
            
            submitted = st.form_submit_button("📝 업데이트 등록", use_container_width=True, type="primary")
            
            if submitted:
                success = self._add_process_event(
                    order_id=order_id,
                    item=None if selected_item == "전체 품목" else selected_item,
                    stage=stage,
                    done_at=done_date,
                    progress=progress,
                    note=note
                )
                
                if success:
                    st.success("✅ 진행 상황이 업데이트되었습니다!")
                    st.rerun()
                else:
                    st.error("❌ 업데이트 중 오류가 발생했습니다.")

    def _add_process_event(self, order_id, item, stage, done_at, progress, note):
        """새로운 프로세스 이벤트 추가"""
        try:
            # 현재 이벤트 데이터 로드
            _, _, events = self._load_wip_data()
            if events is None:
                events = pd.DataFrame(columns=[
                    "event_id", "order_id", "item", "stage", 
                    "planned_at", "done_at", "progress", "note", 
                    "created_at", "created_by"
                ])
            
            # 새 이벤트 생성
            new_event = {
                "event_id": f"{order_id}_{int(datetime.now().timestamp())}",
                "order_id": str(order_id),
                "item": str(item) if item else "",
                "stage": stage,
                "planned_at": pd.NaT,
                "done_at": pd.to_datetime(done_at),
                "progress": int(progress),
                "note": str(note) if note else "",
                "created_at": pd.Timestamp.now(),
                "created_by": "USER"  # 추후 사용자 관리 시 개선
            }
            
            # 새 이벤트 추가
            new_events = pd.concat([events, pd.DataFrame([new_event])], ignore_index=True)
            
            # 엑셀 파일에 저장
            self._save_events_to_excel(new_events)
            
            return True
            
        except Exception as e:
            st.error(f"이벤트 저장 중 오류: {e}")
            return False

    def _save_events_to_excel(self, events_df):
        """ProcessEvents를 엑셀 파일에 저장"""
        try:
            # 기존 워크북 로드
            wb = load_workbook(self.database_file)
            
            # ProcessEvents 시트 삭제 후 재생성
            if "ProcessEvents" in wb.sheetnames:
                del wb["ProcessEvents"]
            
            ws = wb.create_sheet("ProcessEvents")
            
            # 헤더 추가
            headers = list(events_df.columns)
            ws.append(headers)
            
            # 데이터 추가
            for _, row in events_df.iterrows():
                row_data = []
                for col in headers:
                    value = row[col]
                    if pd.isna(value):
                        row_data.append("")
                    elif isinstance(value, pd.Timestamp):
                        row_data.append(value.strftime("%Y-%m-%d %H:%M:%S"))
                    else:
                        row_data.append(str(value))
                ws.append(row_data)
            
            # 저장
            wb.save(self.database_file)
            
        except Exception as e:
            raise Exception(f"엑셀 저장 실패: {e}")

    def render_wip_dashboard(self):
        """메인 WIP 대시보드 렌더링"""
        st.title("🏗️ WIP 현황 관리 시스템")
        st.markdown("---")
        
        # 데이터 로딩
        orders, items, events = self._load_wip_data()
        
        if orders is None:
            st.error("데이터를 로드할 수 없습니다.")
            return
        
        # 데이터가 비어있는 경우
        if orders.empty:
            st.info("📝 발주 데이터가 없습니다. Orders 시트에 데이터를 추가해주세요.")
            st.markdown("**필요한 컬럼**: order_id, project, vendor, order_date, due_date, status")
            return
        
        # 진행률 계산
        orders = self._calculate_progress(orders, events)
        
        # 대시보드 카드
        self._render_dashboard_cards(orders)
        
        st.markdown("---")
        
        # 필터 적용
        filters = self._render_filters(orders)
        filtered_orders = self._apply_filters(orders, filters)
        
        st.markdown("---")
        
        # 메인 테이블
        selected_order_id = self._render_main_table(filtered_orders, items)
        
        # 선택된 발주의 상세 정보
        if selected_order_id:
            st.markdown("---")
            self._render_detail_panel(selected_order_id, orders, items, events)
            
            st.markdown("---")
            self._render_update_form(selected_order_id, items)

# 메인 실행 부분
def main():
    """메인 함수"""
    wip_manager = WIPManager()
    wip_manager.render_wip_dashboard()

if __name__ == "__main__":
    main()